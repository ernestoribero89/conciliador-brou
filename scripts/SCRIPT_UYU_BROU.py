import pandas as pd, re, sys, unicodedata
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

TOLERANCIA = 0.9
DIAS = 3
VERDE        = PatternFill('solid', fgColor='92D051')
AMARILLO     = PatternFill('solid', fgColor='FFFF00')
ROSA         = PatternFill('solid', fgColor='FF99CC')
GRIS         = PatternFill('solid', fgColor='C0C0C0')

def nan_safe(v): return v is None or (isinstance(v, float) and pd.isna(v))
def to_float(v):
    try: return float(v)
    except: return None
def norm(s): return unicodedata.normalize('NFD', s).encode('ascii','ignore').decode('ascii').lower()
def extract_wiz(s):
    if not isinstance(s, str): return None
    m = re.search(r'(Wiz\d{8}n\d+)', s, re.IGNORECASE)
    return m.group(1) if m else None
def is_comision(d):
    if not isinstance(d, str): return False
    n = norm(d)
    return 'comision' in n or 'apertura coe' in n
def is_dif_cambio(c):
    if not isinstance(c, str): return False
    cl = c.lower()
    return ('diferencia' in cl or 'dif' in cl) and ('cambio' in cl or 'tipo' in cl)
def is_sueldo_sap(c):
    if not isinstance(c, str): return False
    n = norm(c)
    return any(x in n for x in [
        'sueldo', 'sueldos',
        'salario', 'salarios',
        'adel rem', 'adelanto rem', 'adelanto remuneracion',
        'adelanto', 'haberes'
    ])

def is_sueldo_bco(d):
    if not isinstance(d, str): return False
    n = norm(d)
    return any(x in n for x in [
        'sueldo', 'sueldos',
        'salario', 'salarios',
        'pago de salarios',
        'pago ch. ventanilla',
        'trf e-brou sueldos',
        'trf spi sueldos',
        'haberes'
    ])

def is_no_sueldo_bco(d):
    if not isinstance(d, str): return False
    n = norm(d).upper()
    return (
        'PAGO A PROVEEDORES' in n
        or 'PAGO PROV' in n
        or 'PROVEEDOR' in n
        or 'COMISION' in n
        or 'APERTURA COE' in n
    )

def is_trf_sueldo_compatible(d):
    if not isinstance(d, str): return False
    n = norm(d)
    du = n.upper()
    return (
        is_sueldo_bco(d)
        or (
            du.startswith('TRF')
            and 'PAGO A' not in du
            and 'PAGO PROV' not in du
            and 'PROVEEDOR' not in du
            and 'COMISION' not in du
        )
    )
def is_ente_sap(c):
    if not isinstance(c, str): return False
    return any(x in c.upper() for x in ['ANTEL','BPS','BSE','UTE','OSE','MGAP'])
def wiz_date(w):
    m = re.search(r'Wiz(\d{4})(\d{2})(\d{2})', w, re.IGNORECASE)
    if m: return pd.Timestamp(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    return None

SAP_PATH = sys.argv[1] if len(sys.argv) > 1 else '/mnt/user-data/uploads/TEST_BROU_ABRIL_SAP.xlsx'
BCO_PATH = sys.argv[2] if len(sys.argv) > 2 else '/mnt/user-data/uploads/BROU_UYU_ABRIL_TEST.xlsx'
OUT_SAP  = sys.argv[3] if len(sys.argv) > 3 else '/mnt/user-data/outputs/OUT_SAP_UYU.xlsx'
OUT_BCO  = sys.argv[4] if len(sys.argv) > 4 else '/mnt/user-data/outputs/OUT_BANCO_UYU.xlsx'

# ── Carga ──────────────────────────────────────────────────────────────────────
df_sap = pd.read_excel(SAP_PATH, header=None, dtype=str)
sap = []
for i, row in df_sap.iterrows():
    if i < 2: continue
    f = pd.to_datetime(row[0], dayfirst=True, errors='coerce')
    if pd.isna(f): continue
    com = str(row[6]).strip() if not nan_safe(row[6]) else ''
    imp = to_float(row[33])
    # Diferencias de cambio → gris, sin cruce
    if is_dif_cambio(com):
        sap.append(dict(er=i+1, f=f, doc=str(row[2]).strip() if not nan_safe(row[2]) else '',
                        com=com, imp=imp, wiz=None, m=False, an=True, color=GRIS))
        continue
    if imp is None: continue
    sap.append(dict(er=i+1, f=f, doc=str(row[2]).strip() if not nan_safe(row[2]) else '',
                    com=com, imp=imp, wiz=extract_wiz(com), m=False, an=False, color=VERDE))

df_b = pd.read_excel(BCO_PATH, header=None, dtype=str)
bco = []
for i, row in df_b.iterrows():
    if i < 14: continue
    f = pd.to_datetime(row[0], dayfirst=True, errors='coerce')
    if pd.isna(f): continue
    desc = str(row[1]).strip() if not nan_safe(row[1]) else ''
    if 'saldo' in desc.lower(): continue
    deb = to_float(row[7]); cre = to_float(row[8])
    if deb and deb > 0: imp = -deb
    elif cre and cre > 0: imp = cre
    else: continue
    bco.append(dict(er=i+1, f=f, desc=desc, imp=imp, m=False, color=VERDE))

# ── Anulados SAP ───────────────────────────────────────────────────────────────
for r in sap:
    if r['an']: continue
    cl = r['com'].lower()
    if ('cancelar' in cl or 'anular' in cl) and 'entrada para' in cl:
        r['an'] = True
        m2 = re.search(r'pago\s+\[?recibido\]?\s+(\d+)', cl)
        if not m2: m2 = re.search(r'(\d{4,6})\s*$', r['com'].strip())
        if m2:
            num = m2.group(1)
            for r2 in sap:
                if not r2['an'] and re.sub(r'\D','', r2['doc']) == num:
                    r2['an'] = True

# ── Wiz individual (N SAP → 1 banco) ──────────────────────────────────────────
wg = defaultdict(list)
for si, sr in enumerate(sap):
    if sr['an'] or sr['m']: continue
    w = sr['wiz']
    if w and not nan_safe(w): wg[w].append(si)

wiz_matched = set()
for w, idxs in wg.items():
    total = sum(sap[i]['imp'] for i in idxs)
    fs = wiz_date(w) or sap[idxs[0]]['f']
    for bi, br in enumerate(bco):
        if br['m']: continue
        if abs(abs(br['imp']) - abs(total)) <= TOLERANCIA and abs((br['f'] - fs).days) <= DIAS:
            for si in idxs: sap[si]['m'] = True
            bco[bi]['m'] = True
            wiz_matched.add(w)
            break

# ── Wiz combinado: múltiples grupos sin match de misma fecha → 1 banco ────────
wiz_unmatched = {w: idxs for w, idxs in wg.items()
                 if w not in wiz_matched and not sap[idxs[0]]['m']}
by_wiz_date = defaultdict(list)
for w, idxs in wiz_unmatched.items():
    fd = wiz_date(w)
    if fd: by_wiz_date[fd].append((w, idxs))

for fd, group in by_wiz_date.items():
    all_idxs = [i for _, idxs in group for i in idxs if not sap[i]['m']]
    if not all_idxs: continue
    total = sum(sap[i]['imp'] for i in all_idxs)
    for bi, br in enumerate(bco):
        if br['m']: continue
        if abs(abs(br['imp']) - abs(total)) <= TOLERANCIA and abs((br['f'] - fd).days) <= DIAS:
            for i in all_idxs:
                sap[i]['m'] = True
                sap[i]['color'] = VERDE
            bco[bi]['m'] = True
            bco[bi]['color'] = VERDE
            break

# ── Wiz + anulado + reemisión → verde oscuro ──────────────────────────────────
# Caso: un PP del grupo Wiz fue anulado (hay fila cancelación con mismo doc).
# El banco pagó: (suma Wiz - PP_anulado) + PP_reemitido_suelto
# Identificar el anulado: buscar en sap filas con an=True cuyo doc referenciado
# pertenece al grupo Wiz.

wg_full = defaultdict(list)
for si, sr in enumerate(sap):
    w = sr['wiz']
    if w and not nan_safe(w): wg_full[w].append(si)

for w, idxs in wg_full.items():
    if all(sap[i]['m'] for i in idxs if not sap[i]['an']): continue
    fecha_wiz = wiz_date(w) or sap[idxs[0]]['f']

    # Buscar filas anuladas cuyo doc referenciado está en este grupo Wiz
    docs_en_grupo = {re.sub(r'\D','', sap[i]['doc']) for i in idxs}
    anulados_ref = []  # (si_cancelacion, si_original)
    for si_c, sr_c in enumerate(sap):
        if not sr_c['an']: continue
        cl = sr_c['com'].lower()
        m2 = re.search(r'(\d{4,6})\s*$', sr_c['com'].strip())
        if not m2: continue
        ref_num = m2.group(1)
        if ref_num in docs_en_grupo:
            # encontrar el original en el grupo
            for si_o in idxs:
                if re.sub(r'\D','', sap[si_o]['doc']) == ref_num:
                    anulados_ref.append((si_c, si_o))
                    break

    if not anulados_ref: continue

    # Suma Wiz sin los originales anulados
    anulados_orig_idxs = {si_o for _, si_o in anulados_ref}
    total_wiz_neto = sum(sap[i]['imp'] for i in idxs if i not in anulados_orig_idxs)

    # Buscar PP sueltos (sin Wiz, no anulados, no matcheados) cerca de fecha_wiz
    suelta_cands = [si for si, sr in enumerate(sap)
                    if not sr['m'] and not sr['an']
                    and (sr['wiz'] is None or nan_safe(sr['wiz']))
                    and abs((sr['f'] - fecha_wiz).days) <= DIAS]

    for si2 in suelta_cands:
        total_real = total_wiz_neto + sap[si2]['imp']
        for bi, br in enumerate(bco):
            if br['m']: continue
            if abs(abs(br['imp']) - abs(total_real)) <= TOLERANCIA and abs((br['f'] - fecha_wiz).days) <= DIAS:
                # Match: colorear solo movimientos vigentes.
                # Los anulados ya detectados quedan bloqueados en GRIS y no vuelven a VERDE.
                for i in idxs:
                    if sap[i]['an']:
                        sap[i]['color'] = GRIS
                        continue
                    sap[i]['m'] = True
                    sap[i]['color'] = VERDE
                for si_c, _ in anulados_ref:
                    sap[si_c]['an'] = True
                    sap[si_c]['m'] = False
                    sap[si_c]['color'] = GRIS
                sap[si2]['m'] = True
                sap[si2]['color'] = VERDE
                bco[bi]['m'] = True
                bco[bi]['color'] = VERDE
                break
        else:
            continue
        break


# ── Sueldos UYU: 1 SAP → N Banco sin nemotécnico claro ────────────────────────
# Regla:
#   - Solo aplica a UYU.
#   - SAP se identifica por comentario sueldo/salario/adelanto/haberes.
#   - Banco se identifica por sueldo/salario/pago ch. ventanilla/TRF compatible.
#   - Nunca mezcla proveedores ni comisiones.
#   - Busca primero match individual y luego suma de banco por fecha ±3 días.
#   - Si la suma no cierra dentro de tolerancia, no fuerza match.

for sr in sap:
    if sr['m'] or sr['an']: continue
    if not is_sueldo_sap(sr['com']): continue

    imp_s = abs(sr['imp'])
    matched = False

    # 1) Match individual exacto / tolerancia contra banco sueldo compatible.
    for br in bco:
        if br['m']: continue
        if br['imp'] >= 0: continue
        if is_no_sueldo_bco(br['desc']): continue
        if not is_trf_sueldo_compatible(br['desc']): continue
        if abs((br['f'] - sr['f']).days) > DIAS: continue

        if abs(abs(br['imp']) - imp_s) <= TOLERANCIA:
            sr['m'] = True
            sr['color'] = VERDE
            br['m'] = True
            br['color'] = VERDE
            matched = True
            break

    if matched: continue

    # 2) Match por suma: candidatos sueldo / TRF compatible por fecha cercana.
    #    Prueba fecha exacta primero, luego +1, -1, +2, -2, +3, -3.
    for dias_offset in range(0, DIAS + 1):
        for sign in ([0] if dias_offset == 0 else [1, -1]):
            target_date = (sr['f'] + pd.Timedelta(days=dias_offset * sign)).date()

            cands = [
                bi for bi, br in enumerate(bco)
                if not br['m']
                and br['imp'] < 0
                and br['f'].date() == target_date
                and not is_comision(br['desc'])
                and not is_no_sueldo_bco(br['desc'])
                and is_trf_sueldo_compatible(br['desc'])
            ]

            if not cands:
                continue

            total = sum(abs(bco[bi]['imp']) for bi in cands)

            if abs(total - imp_s) <= TOLERANCIA:
                sr['m'] = True
                sr['color'] = VERDE
                for bi in cands:
                    bco[bi]['m'] = True
                    bco[bi]['color'] = VERDE
                matched = True
                break

        if matched:
            break

# ── Entes ──────────────────────────────────────────────────────────────────────
entes = ['ANTEL','BPS','BSE','UTE','OSE','MGAP']
for sr in sap:
    if sr['m'] or sr['an']: continue
    if not is_ente_sap(sr['com']): continue
    cu = sr['com'].upper()
    ente = next((x for x in entes if x in cu), None)
    if not ente: continue
    imp_s = abs(sr['imp'])
    cands = [bi for bi, br in enumerate(bco)
             if not br['m'] and abs((br['f'] - sr['f']).days) <= DIAS
             and ente in br['desc'].upper() and br['imp'] < 0]
    if not cands: continue
    total = sum(abs(bco[bi]['imp']) for bi in cands)
    if abs(total - imp_s) <= TOLERANCIA:
        sr['m'] = True
        for bi in cands: bco[bi]['m'] = True

# ── Inverso: N banco misma desc+fecha → 1 SAP ─────────────────────────────────
gb = defaultdict(list)
for bi, br in enumerate(bco):
    if br['m']: continue
    gb[(br['f'].date(), br['desc'])].append(bi)

for sr in sap:
    if sr['m'] or sr['an']: continue
    for (fd, d2), idxs in list(gb.items()):
        if abs((fd - sr['f'].date()).days) > DIAS: continue
        unm = [bi for bi in idxs if not bco[bi]['m']]
        if not unm: continue
        if abs(sum(abs(bco[bi]['imp']) for bi in unm) - abs(sr['imp'])) <= TOLERANCIA:
            sr['m'] = True
            for bi in unm: bco[bi]['m'] = True
            break

# ── Simple ─────────────────────────────────────────────────────────────────────
for sr in sap:
    if sr['m'] or sr['an']: continue
    for br in bco:
        if br['m']: continue
        if abs(abs(br['imp']) - abs(sr['imp'])) <= TOLERANCIA and abs((br['f'] - sr['f']).days) <= DIAS:
            sr['m'] = True; br['m'] = True; break

# ── Stats ──────────────────────────────────────────────────────────────────────
val = [r for r in sap if not r['an']]
sm = sum(1 for r in val if r['m'])
bm = sum(1 for r in bco if r['m'])
comis = [r for r in bco if is_comision(r['desc'])]
print(f"=== BROU UYU ===")
print(f"SAP    : {sm}/{len(val)} matcheados  ({sum(1 for r in sap if r['an'])} excluidos/anulados)")
print(f"Banco  : {bm}/{len(bco)} matcheados")
print(f"Comisiones: {sum(abs(r['imp']) for r in comis):.2f} UYU")
print(f"\nSAP sin match ({len(val)-sm}):")
for r in val:
    if not r['m']: print(f"  {r['f'].date()} | {r['doc']} | {r['com'][:55]} | {r['imp']}")
print(f"\nBanco sin match ({len(bco)-bm-len(comis)}):")
for r in bco:
    if not r['m'] and not is_comision(r['desc']): print(f"  {r['f'].date()} | {r['desc'][:55]} | {r['imp']}")

# ── Escribir SAP output: solo 4 columnas + color en col D ─────────────────────
# Columnas originales a conservar: 0=Fecha, 2=Nº Doc, 6=Comentarios, 33=Sdo Vencido ME
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

wb_orig = load_workbook(SAP_PATH)
ws_orig = wb_orig.active
wb = Workbook()
ws = wb.active

KEEP_COLS = [1, 3, 7, 34]  # 1-based: Fecha, Nº Doc, Comentarios, Sdo Vencido ME
# Copiar fila 1 y 2 (headers y totales) y luego datos
for src_row in ws_orig.iter_rows():
    row_idx = src_row[0].row
    new_row = []
    for orig_col in KEEP_COLS:
        cell = ws_orig.cell(row=row_idx, column=orig_col)
        new_row.append(cell.value)
    ws.append(new_row)
    # Aplicar formato fecha en col A (col 1)
    if row_idx >= 3:
        ws.cell(row=row_idx, column=1).number_format = 'DD/MM/YYYY'

# Aplicar colores en col D (columna 4 = Sdo Vencido ME)
for r in sap:
    c = ws.cell(row=r['er'], column=4)
    if r['an']:  c.fill = GRIS
    elif r['m']: c.fill = r['color']
    else:        c.fill = ROSA

# Ancho de columnas
ws.column_dimensions['A'].width = 14
ws.column_dimensions['B'].width = 14
ws.column_dimensions['C'].width = 55
ws.column_dimensions['D'].width = 18
wb.save(OUT_SAP)

# ── Colorear Banco ─────────────────────────────────────────────────────────────
wb2 = load_workbook(BCO_PATH)
ws2 = wb2.active
tc = []
for r in bco:
    col = 8 if r['imp'] < 0 else 9
    c = ws2.cell(row=r['er'], column=col)
    if is_comision(r['desc']): c.fill = AMARILLO; tc.append(abs(r['imp']))
    elif r['m']:               c.fill = r['color']
    else:                      c.fill = ROSA
if tc:
    lr = max(r['er'] for r in bco) + 2
    ws2.cell(row=lr, column=1, value='TOTAL COMISIONES')
    cx = ws2.cell(row=lr, column=8, value=round(sum(tc), 2))
    cx.fill = AMARILLO
wb2.save(OUT_BCO)
print(f"\nArchivos guardados: {OUT_SAP} | {OUT_BCO}")
